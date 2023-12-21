from openpyxl import load_workbook

class Excel_Functions:
    def __init__(self, excel_file_name, excel_sheet_name):
       self.file = excel_file_name
       self.sheet = excel_sheet_name
       


    # fetch the row count
    def row_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_row


    # fetch the column count
    def column_count(self):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.max_column


    # read data from excel file
    def read_data(self, row_number, column_number):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       return sheet.cell(row=row_number, column=column_number).value


    # write data into excel file
    def write_data(self, row_number, column_number, data):
       workbook = load_workbook(self.file)
       sheet = workbook[self.sheet]
       sheet.cell(row=row_number, column=column_number).value = data
       workbook.save(self.file)

    # Gives details to the test_login method in Test_login page for the purpose of running pytest
    def read_excel_login_test_detail_util(self):
        test_details = []
        
        for row in range(2,4):
            test_detail_tuple = ()
            username = self.read_data(row, 9)
            password = self.read_data(row, 10)
            exp_result = self.read_data(row, 12)
            test_detail_tuple = (username,password,row,exp_result)
            test_details.append(test_detail_tuple)

        print("Test Data login test : ",test_details)
        return test_details

    def fetch_test_details_pim_01(self):
        try:
            test_data_lst = []
            
            for row_num in range(2,3):
                # Sheet2
                # Login username and password from 
                test_data_tuple = ()
                lg_username = self.read_data(row_num,1)
                lg_password = self.read_data(row_num,2)
                row = self.read_data(row_num,3)
                exp_res = self.read_data(row_num,4)
                
                #Employee details
                emp_id_da = self.read_data(row_num,5)
                employee_fname_da = self.read_data(row_num,6)
                employee_lname_da = self.read_data(row_num,7)
                #employee_id_da = self.excel_pim_01.read_data(row_num,3)
                employee_user_name = self.read_data(row_num,8)
                employee_password = self.read_data(row_num,9)

                # Employee personal details
                nick_name_da = self.read_data(row_num,10)
                other_id_da = self.read_data(row_num,11)
                driver_lic_da = self.read_data(row_num,12)
                lic_exp_da = self.read_data(row_num,13)
                ssn_da = self.read_data(row_num,14)
                sin_da = self.read_data(row_num,15)
                dob_da = self.read_data(row_num,16)
                mili_ser_da =self.read_data(row_num,17)



                test_data_tuple = (lg_username,lg_password,row,exp_res,emp_id_da,employee_fname_da,employee_lname_da
                                ,employee_user_name,employee_password,nick_name_da,other_id_da,driver_lic_da,lic_exp_da,
                                ssn_da,sin_da,dob_da,mili_ser_da
                                )
                #print(test_data_tuple)
                test_data_lst.append(test_data_tuple)

            print("Test data pim_01: ",test_data_lst)
            return test_data_lst
        
        except Exception as e:
            print("Fetching test datas from excel for pim 01 failed : ",e)
            return []
    
    def fetch_test_details_pim_02(self):
        try:
            test_data_lst = []
            
            for row_num in range(2,3):
                # Sheet3
                # Login username and password from 
                test_data_tuple = ()
                lg_username = self.read_data(row_num,1)
                lg_password = self.read_data(row_num,2)
                row = self.read_data(row_num,3)
                exp_res = self.read_data(row_num,4)
                
                #Employee details 
                emp_id_da = self.read_data(row_num,5)
                employee_fname_da = self.read_data(row_num,6)
                employee_lname_da = self.read_data(row_num,7)
                #employee_id_da = self.excel_pim_01.read_data(row_num,3)
                employee_user_name = self.read_data(row_num,8)
                employee_password = self.read_data(row_num,9)

                # Employee personal details
                nick_name_da = self.read_data(row_num,10)
                other_id_da = self.read_data(row_num,11)
                driver_lic_da = self.read_data(row_num,12)
                lic_exp_da = self.read_data(row_num,13)
                ssn_da = self.read_data(row_num,14)
                sin_da = self.read_data(row_num,15)
                dob_da = self.read_data(row_num,16)
                mili_ser_da =self.read_data(row_num,17)
                edited_fname_da =self.read_data(row_num,18)
                edited_nationality_da = self.read_data(row_num,19)
                edited_blood_type_da = self.read_data(row_num,20)



                test_data_tuple = (lg_username,lg_password,row,exp_res,emp_id_da,employee_fname_da,employee_lname_da,
                                employee_user_name,employee_password,nick_name_da,other_id_da,driver_lic_da,lic_exp_da,
                                ssn_da,sin_da,dob_da,mili_ser_da,edited_fname_da,edited_nationality_da,edited_blood_type_da
                                )
                #print(test_data_tuple)
                test_data_lst.append(test_data_tuple)

            print("Test data pim_02: ",test_data_lst)
            return test_data_lst
        
        except Exception as e:
            print("Fetching test datas from excel for pim 02 failed : ",e)
            return []
    
    def fetch_test_details_pim_03(self):
        try:
            test_data_lst = []
            
            for row_num in range(2,3):
                # Sheet4
                # Login username and password from 
                test_data_tuple = ()
                lg_username = self.read_data(row_num,1)
                lg_password = self.read_data(row_num,2)
                row = self.read_data(row_num,3)
                exp_res = self.read_data(row_num,4)
                
                #Employee details 
                emp_id_da = self.read_data(row_num,5)
                employee_fname_da = self.read_data(row_num,6)
                employee_lname_da = self.read_data(row_num,7)
                employee_user_name = self.read_data(row_num,8)
                employee_password = self.read_data(row_num,9)

                # Employee personal details
                nick_name_da = self.read_data(row_num,10)
                other_id_da = self.read_data(row_num,11)
                driver_lic_da = self.read_data(row_num,12)
                lic_exp_da = self.read_data(row_num,13)
                ssn_da = self.read_data(row_num,14)
                sin_da = self.read_data(row_num,15)
                dob_da = self.read_data(row_num,16)
                mili_ser_da =self.read_data(row_num,17)


                test_data_tuple = (lg_username,lg_password,row,exp_res,emp_id_da,employee_fname_da,employee_lname_da,
                                employee_user_name,employee_password,nick_name_da,other_id_da,driver_lic_da,lic_exp_da,
                                ssn_da,sin_da,dob_da,mili_ser_da,
                                )
                #print(test_data_tuple)
                test_data_lst.append(test_data_tuple)

            print("Test data pim_03 : ",test_data_lst)
            return test_data_lst
        
        except Exception as e:
            print("Fetching test datas from excel for pim 02 failed : ",e)
            return []
    
'''ef = Excel_Functions("Capstone_1.xlsx","Sheet5")
ef.fetch_test_details_pim_01()
ef.fetch_test_details_pim_02()
ef.fetch_test_details_pim_03()'''


'''class Base_driver_cls:
    excel_login = Excel_Functions("Capstone_1.xlsx", "Sheet1")
    excel_pim_01 = Excel_Functions("Capstone_1.xlsx", "Sheet2")

    @classmethod
    def fetch_test_details_pim_01(cls):
        try:
            test_data_lst = []
            
            for row_num in range(2,3):
                # Sheet1
                # Login username and password from 
                test_data_tuple = ()
                result_row = 4
                lg_username = cls.excel_login.read_data(result_row,9)
                lg_password = cls.excel_login.read_data(result_row,10)
                exp_res = cls.excel_login.read_data(result_row,12)
                
                # Sheet2
                #Employee details 
                employee_fname_da = cls.excel_pim_01.read_data(row_num,1)
                employee_lname_da = cls.excel_pim_01.read_data(row_num,2)
                #employee_id_da = self.excel_pim_01.read_data(row_num,3)
                employee_user_name = cls.excel_pim_01.read_data(row_num,3)
                employee_password = cls.excel_pim_01.read_data(row_num,4)

                # Employee personal details
                nick_name_da = cls.excel_pim_01.read_data(row_num,5)
                other_id_da = cls.excel_pim_01.read_data(row_num,6)
                driver_lic_da = cls.excel_pim_01.read_data(row_num,7)
                lic_exp_da = cls.excel_pim_01.read_data(row_num,8)
                ssn_da = cls.excel_pim_01.read_data(row_num,9)
                sin_da = cls.excel_pim_01.read_data(row_num,10)
                dob_da = cls.excel_pim_01.read_data(row_num,11)
                mili_ser_da =cls.excel_pim_01.read_data(row_num,12)



                test_data_tuple = (lg_username,lg_password,result_row,
                                exp_res,employee_fname_da,employee_lname_da
                                ,employee_user_name,employee_password,
                                nick_name_da,other_id_da,driver_lic_da,lic_exp_da,
                                ssn_da,sin_da,dob_da,mili_ser_da
                                )
                print(test_data_tuple)
                test_data_lst.append(test_data_tuple)

            print("Test data : ",test_data_lst)
            return test_data_lst
        
        except Exception as e:
            print("Fetching test datas from excel for pim 01 failed : ",e)
            return []'''
    
    

    
    

